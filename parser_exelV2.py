import xlsxwriter
import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile
import xlrd
from openpyxl import load_workbook

dc={}
last_date={}
not_write = {}

def main_logic():

    global dc
    global not_write
    t_doc12 = []
    t_doc2 = []
    t_doc4 = []
    t_doc5 = []
    t_doc7 = []
    t_doc8 = []
    t_doc9 = []

    key2 = []
    doc12 = []
    doc2 = []
    doc4 = []
    doc5 = []
    doc7 = []
    doc8 = []
    doc9 = []

    key1 = []
    key = []

    wb = load_workbook(r'patch to exel')
    ws = wb.active
    sheet = wb['sheet1']
    name = []
    dc = {}
    hello = []

    for row in sheet.rows:
        dc[row[0].value] = row[1].value
        key1.append(row[1].value)
        t_doc12.append(row[4].value)
        t_doc2.append(row[5].value)
        t_doc4.append(row[6].value)
        t_doc5.append(row[7].value)
        t_doc7.append(row[8].value)
        t_doc8.append(row[9].value)
        t_doc9.append(row[10].value)
        name.append(row[0].value)

    for k in key1:
        if k == None:
            key.append(None)
        if k == str(k):
            key.append(k.replace(' ', ''))
    print(key)

    for k in t_doc12:
        if k == None:
            doc12.append(None)
        if k == str(k):
            doc12.append(k.replace(' ', ''))
    print(doc12)

    for k in t_doc2:
        if k == None:
            doc2.append(None)
        if k == str(k):
            doc2.append(k.replace(' ', ''))
    print(doc2)

    for k in t_doc4:
        if k == None:
            doc4.append(None)
        if k == str(k):
            doc4.append(k.replace(' ', ''))
    print(doc4)

    for k in t_doc5:
        if k == None:
            doc5.append(None)
        if k == str(k):
            doc5.append(k.replace(' ', ''))
    print(doc5)

    for k in t_doc7:
        if k == None:
            doc7.append(None)
        if k == str(k):
            doc7.append(k.replace(' ', ''))
    print(doc7)

    for k in t_doc8:
        if k == None:
            doc8.append(None)
        if k == str(k):
            doc8.append(k.replace(' ', ''))
    print(doc8)

    for k in t_doc9:
        if k == None:
            doc9.append(None)
        if k == str(k):
            doc9.append(k.replace(' ', ''))
    print(doc9)

    wb.close()
    workbook = xlsxwriter.Workbook(r'patch to exel')
    worksheet = workbook.add_worksheet()
    print(worksheet)
    row = 0
    row_name = 0
    print(name)
    for n in name:
        worksheet.write(row_name, 0, n)
        row_name += 1
    row = 0
    name = list()

    for element in key:
        worksheet.write(row, 1, element)
        if element not in doc12:
            worksheet.write(row, 4, "")
            # print(element + (" не подписал документ 2"))
            name.append(element)
        if element in doc12:
            worksheet.write(row, 4, element)

        if element not in doc2:
            worksheet.write(row, 5, "")
            # print(element + (" не подписал документ 2"))
            name.append(element)
        if element in doc2:
            worksheet.write(row, 5, element)

        if element not in doc4:
            worksheet.write(row, 6, "")
            # print(element + ("не подписал документ 4"))
            name.append(element)
        if element in doc4:
            worksheet.write(row, 6, element)

        if element not in doc5:
            worksheet.write(row, 7, "")
            # print(element + ("не подписал документ 5"))
            name.append(element)
        if element in doc5:
            worksheet.write(row, 7, element)

        if element not in doc7:
            # print(element + ("не подписал документ 7"))
            name.append(element)
            worksheet.write(row, 8, "")
        if element in doc7:
            worksheet.write(row, 8, element)

        if element not in doc8:
            # print(element + ("не подписал документ 8"))
            name.append(element)
            worksheet.write(row, 9, "")
        if element in doc8:
            worksheet.write(row, 9, element)

        if element not in doc9:
            worksheet.write(row, 10, "")
            # print(element + ("не подписал документ 9"))
            name.append(element)
        if element in doc9:
            worksheet.write(row, 10, element)

        row += 1
    print(name)
    workbook.close()

    string_task1 = name
    count_task1 = {}

    workbook = xlsxwriter.Workbook(r'patch to exel')
    worksheet = workbook.add_worksheet()

    for s in string_task1:
        if s in count_task1:
            count_task1[s] += 1
        else:
            count_task1[s] = 1

    row1 = 0
    for key in count_task1:
        if count_task1[key] >= 1:
            print(key, end=',')
            key2.append(key)
            worksheet.write(row1, 1, key)
            worksheet.write(row1, 2, count_task1[key])

            not_write[key] = count_task1[key]
            row1 += 1
    row2 = 0
    for l in key2:
        if l in dc.values():
            worksheet.write(row2, 0, get_key(dc, l))
            row2 += 1
        if l not in dc.values():
            worksheet.write(row2, 0, " ")
            row2 += 1

    workbook.close()
    comparison_last()

def get_key(d, value):
    for k, v in d.items():
        if v == value:
            return k

def comparison_last():
    global last_date
    global dc
    global not_write
    wb = load_workbook(r'patch to exel')
    ws = wb.active
    sheet = wb['Sheet1']
    for row in sheet.rows:
        last_date[row[1].value] = str(row[2].value)
    print(last_date)
    print(not_write)
    wb.close()

    workbook = xlsxwriter.Workbook(r'patch to exel')
    worksheet = workbook.add_worksheet()
    row = 0
    timpe=[]
    for key, val in last_date.items():
        if key in not_write:
            print(key, str(not_write[key]))
            worksheet.write(row, 1, key)
            worksheet.write(row, 2, str(not_write[key]))
            timpe.append(key)
            if str(val) != str(not_write[key]):
                print('test')
                worksheet.write(row, 3, ''.join(str(not_write[key])))

                print("не совпало", key,''.join(str(not_write[key])))
            row+=1
    row2=0
    for l in timpe:
        if l in dc.values():
            worksheet.write(row2, 0, get_key(dc, l))
            row2 += 1
        if l not in dc.values():
            worksheet.write(row2, 0, " ")
            row2 += 1
    workbook.close()
if __name__ == "__main__":
    main_logic()
